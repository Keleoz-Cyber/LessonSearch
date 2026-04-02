"""
Excel 结构分析器：扫描考勤表目录，自动解析年级/专业/班级/学生信息。

支持两种格式:
  1. 考勤表格式: Row 1-2 标题, Row 3-4 表头, Row 5+ 数据 (B=姓名, C=学号)
     年级从文件夹名提取，专业从文件名提取，班级从 sheet 名或 A1 标题提取
  2. 名单格式: Row 1 表头, Row 2+ 数据，列名自动识别
     班级信息在数据列中（如"计科2401"），年级/专业从班级名解析

用法:
    python excel_analyzer.py                          # 分析默认目录
    python excel_analyzer.py --dir /path/to/data      # 指定目录
    python excel_analyzer.py --json                   # 输出 JSON 格式
    python excel_analyzer.py --verbose                # 显示每个学生的详细数据

该脚本独立可运行，不依赖数据库连接。
"""
import argparse
import json
import os
import re
import sys
from dataclasses import dataclass, field, asdict
from typing import Optional

from openpyxl import load_workbook

from config import EXCEL_DATA_DIR, MAJOR_MAPPING


# ============================================================
# 数据结构
# ============================================================

@dataclass
class StudentRecord:
    name: str
    student_no: str
    class_label: Optional[str]  # 班级标签，如 "计科2401"（名单格式时有值）
    row_number: int


@dataclass
class SheetAnalysis:
    sheet_name: str
    format_type: str                    # "attendance" / "roster"
    class_code: Optional[str]           # 4 位班级编号（考勤表格式时有值）
    class_code_source: Optional[str]
    title_text: Optional[str]
    students: list[StudentRecord] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class FileAnalysis:
    file_path: str
    file_name: str
    folder_name: str
    format_type: str                    # "attendance" / "roster"
    grade_year: Optional[int]
    grade_source: str
    major_short: Optional[str]
    major_full: Optional[str]
    major_source: str
    sheets: list[SheetAnalysis] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)


@dataclass
class AnalysisReport:
    data_dir: str
    total_files: int = 0
    total_sheets: int = 0
    total_students: int = 0
    files: list[FileAnalysis] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


# ============================================================
# 通用解析函数
# ============================================================

HEADER_KEYWORDS_STUDENT_NO = ["学号"]
HEADER_KEYWORDS_NAME = ["姓名"]
HEADER_KEYWORDS_CLASS = ["班级"]


def clean_student_no(raw: str) -> str:
    """清理学号字符串"""
    s = str(raw).strip()
    if "." in s:
        s = s.split(".")[0]
    return s


def detect_sheet_format(ws) -> str:
    """检测 sheet 格式: attendance (考勤表) 或 roster (名单)"""
    a1 = ws.cell(row=1, column=1).value
    if a1 is None:
        return "attendance"
    a1_str = str(a1).strip()
    # 名单格式: 第一行是表头，包含 "学号"/"姓名" 等
    if a1_str in HEADER_KEYWORDS_STUDENT_NO or a1_str in HEADER_KEYWORDS_NAME:
        return "roster"
    # 考勤表格式: A1 是标题（如 "xxx学院xxx班第 周考勤情况表"）或 "序号"
    return "attendance"


def find_columns(ws, header_row: int = 1) -> dict[str, int]:
    """从表头行自动识别列位置，返回 {字段名: 列号}"""
    cols = {}
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=c).value
        if val is None:
            continue
        val_str = str(val).strip()
        if val_str in HEADER_KEYWORDS_STUDENT_NO:
            cols["student_no"] = c
        elif val_str in HEADER_KEYWORDS_NAME:
            cols["name"] = c
        elif val_str in HEADER_KEYWORDS_CLASS:
            cols["class"] = c
    return cols


# ============================================================
# 考勤表格式分析
# ============================================================

def extract_grade_from_folder(folder_name: str) -> Optional[int]:
    """从文件夹名提取年级: '22考勤表' → 2022"""
    match = re.search(r"(\d{2})", folder_name)
    if match:
        return 2000 + int(match.group(1))
    return None


def extract_major_from_filename(filename: str) -> tuple[Optional[str], Optional[str]]:
    """从文件名提取专业: '22电信.xlsx' → ('电信', '电子信息工程')"""
    for short, full in MAJOR_MAPPING.items():
        if short in filename:
            return short, full
    return None, None


def extract_class_code_from_sheet_name(sheet_name: str) -> Optional[str]:
    if re.match(r"^Sheet\d*$", sheet_name, re.IGNORECASE):
        return None
    match = re.search(r"(\d{4})", sheet_name)
    return match.group(1) if match else None


def extract_class_code_from_title(title_text: str) -> Optional[str]:
    if not title_text:
        return None
    match = re.search(r"(\d{4})", title_text)
    return match.group(1) if match else None


def is_footer_row(row_values: list) -> bool:
    footer_keywords = ["日期", "学委签名", "任课教师签名", "签名"]
    text = " ".join(str(v) for v in row_values if v)
    return any(kw in text for kw in footer_keywords)


def analyze_sheet_attendance(ws, sheet_name: str) -> SheetAnalysis:
    """分析考勤表格式的 sheet"""
    analysis = SheetAnalysis(
        sheet_name=sheet_name, format_type="attendance",
        class_code=None, class_code_source=None, title_text=None,
    )

    title_val = ws.cell(row=1, column=1).value
    if title_val:
        analysis.title_text = str(title_val).strip()

    code = extract_class_code_from_sheet_name(sheet_name)
    if code:
        analysis.class_code = code
        analysis.class_code_source = "sheet_name"
    else:
        code = extract_class_code_from_title(analysis.title_text or "")
        if code:
            analysis.class_code = code
            analysis.class_code_source = "title_cell"
        else:
            analysis.warnings.append(
                f"无法从 sheet 名 '{sheet_name}' 或 A1 标题中提取班级编号，待人工确认"
            )

    for row_idx in range(5, ws.max_row + 1):
        row_values = [ws.cell(row=row_idx, column=c).value for c in range(1, ws.max_column + 1)]
        if is_footer_row(row_values):
            break

        name_val = ws.cell(row=row_idx, column=2).value
        no_val = ws.cell(row=row_idx, column=3).value

        if not name_val or not no_val:
            continue

        name_str = str(name_val).strip()
        no_str = clean_student_no(no_val)

        if not name_str or not no_str:
            continue

        if not re.match(r"^\d{10,14}$", no_str):
            analysis.warnings.append(f"第 {row_idx} 行学号格式异常: '{no_str}'（姓名: {name_str}）")

        analysis.students.append(StudentRecord(
            name=name_str, student_no=no_str, class_label=None, row_number=row_idx,
        ))

    if not analysis.students:
        analysis.warnings.append("该 sheet 未提取到任何学生数据")

    return analysis


# ============================================================
# 名单格式分析
# ============================================================

def parse_class_label(class_label: str) -> tuple[Optional[str], Optional[str], Optional[int]]:
    """
    解析班级标签: '计科2401' → (major_short='计科', class_code='2401', grade_year=2024)
    """
    if not class_label:
        return None, None, None

    major_short = None
    for short in MAJOR_MAPPING:
        if short in class_label:
            major_short = short
            break

    match = re.search(r"(\d{4})", class_label)
    class_code = match.group(1) if match else None

    grade_year = None
    if class_code:
        year_prefix = int(class_code[:2])
        grade_year = 2000 + year_prefix

    return major_short, class_code, grade_year


def analyze_sheet_roster(ws, sheet_name: str) -> SheetAnalysis:
    """分析名单格式的 sheet"""
    analysis = SheetAnalysis(
        sheet_name=sheet_name, format_type="roster",
        class_code=None, class_code_source="data_column", title_text=None,
    )

    cols = find_columns(ws, header_row=1)

    if "student_no" not in cols or "name" not in cols:
        analysis.warnings.append(
            f"无法在表头中识别 '学号' 和 '姓名' 列，找到的列: {cols}"
        )
        return analysis

    no_col = cols["student_no"]
    name_col = cols["name"]
    class_col = cols.get("class")

    for row_idx in range(2, ws.max_row + 1):
        no_val = ws.cell(row=row_idx, column=no_col).value
        name_val = ws.cell(row=row_idx, column=name_col).value

        if not no_val or not name_val:
            continue

        no_str = clean_student_no(no_val)
        name_str = str(name_val).strip()

        if not name_str or not no_str:
            continue

        if not re.match(r"^\d{10,14}$", no_str):
            analysis.warnings.append(f"第 {row_idx} 行学号格式异常: '{no_str}'（姓名: {name_str}）")
            continue

        class_label = None
        if class_col:
            cv = ws.cell(row=row_idx, column=class_col).value
            if cv:
                class_label = str(cv).strip()

        analysis.students.append(StudentRecord(
            name=name_str, student_no=no_str, class_label=class_label, row_number=row_idx,
        ))

    if not analysis.students:
        analysis.warnings.append("该 sheet 未提取到任何学生数据")

    return analysis


# ============================================================
# 文件分析
# ============================================================

def analyze_file(file_path: str) -> FileAnalysis:
    """分析单个 Excel 文件，自动判断格式"""
    file_name = os.path.basename(file_path)
    folder_name = os.path.basename(os.path.dirname(file_path))

    grade_year = extract_grade_from_folder(folder_name)
    major_short, major_full = extract_major_from_filename(file_name)

    # 先探测格式
    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        first_sheet = wb[wb.sheetnames[0]]
        fmt = detect_sheet_format(first_sheet)
        wb.close()
    except Exception:
        fmt = "attendance"

    analysis = FileAnalysis(
        file_path=file_path,
        file_name=file_name,
        folder_name=folder_name,
        format_type=fmt,
        grade_year=grade_year,
        grade_source="folder_name" if grade_year else "unknown",
        major_short=major_short,
        major_full=major_full,
        major_source="file_name" if major_short else "unknown",
    )

    if fmt == "attendance":
        if not grade_year:
            analysis.warnings.append(f"无法从文件夹名 '{folder_name}' 提取年级，待人工确认")
        if not major_short:
            analysis.warnings.append(f"无法从文件名 '{file_name}' 提取专业，待人工确认")

    try:
        wb = load_workbook(file_path, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            ws = wb[sn]
            sheet_fmt = detect_sheet_format(ws)
            if sheet_fmt == "roster":
                sheet_analysis = analyze_sheet_roster(ws, sn)
            else:
                sheet_analysis = analyze_sheet_attendance(ws, sn)
            analysis.sheets.append(sheet_analysis)
        wb.close()
    except Exception as e:
        analysis.warnings.append(f"读取文件出错: {e}")

    # 名单格式: 从数据中提取年级/专业信息
    if fmt == "roster":
        analysis.grade_source = "data_column"
        analysis.major_source = "data_column"
        # 收集所有出现的班级标签以推断年级和专业
        all_class_labels = set()
        for sa in analysis.sheets:
            for stu in sa.students:
                if stu.class_label:
                    all_class_labels.add(stu.class_label)
        if all_class_labels:
            grades_found = set()
            majors_found = set()
            for cl in all_class_labels:
                ms, cc, gy = parse_class_label(cl)
                if gy:
                    grades_found.add(gy)
                if ms:
                    majors_found.add(ms)
            # 如果只有一个年级/专业，可以设置文件级别
            if len(grades_found) == 1:
                analysis.grade_year = grades_found.pop()
            if len(majors_found) == 1:
                ms = majors_found.pop()
                analysis.major_short = ms
                analysis.major_full = MAJOR_MAPPING.get(ms)

    return analysis


# ============================================================
# 目录扫描
# ============================================================

def scan_and_analyze(data_dir: str) -> AnalysisReport:
    """扫描目录下所有 Excel 文件并分析"""
    report = AnalysisReport(data_dir=data_dir)

    if not os.path.isdir(data_dir):
        report.errors.append(f"目录不存在: {data_dir}")
        return report

    excel_files = []
    for root, _dirs, files in os.walk(data_dir):
        for f in files:
            if f.endswith((".xlsx", ".xls")) and not f.startswith("~$"):
                excel_files.append(os.path.join(root, f))

    excel_files.sort()
    report.total_files = len(excel_files)

    if not excel_files:
        report.errors.append(f"目录下未找到 Excel 文件: {data_dir}")
        return report

    for fp in excel_files:
        file_analysis = analyze_file(fp)
        report.files.append(file_analysis)
        report.total_sheets += len(file_analysis.sheets)
        for sheet in file_analysis.sheets:
            report.total_students += len(sheet.students)
        report.warnings.extend(file_analysis.warnings)
        for sheet in file_analysis.sheets:
            report.warnings.extend(
                f"[{file_analysis.file_name} / {sheet.sheet_name}] {w}"
                for w in sheet.warnings
            )

    # 全局学号重复检查
    seen = {}
    for fa in report.files:
        for sa in fa.sheets:
            for stu in sa.students:
                if stu.student_no in seen:
                    prev = seen[stu.student_no]
                    report.warnings.append(
                        f"学号重复: {stu.student_no}（{stu.name} in {fa.file_name}/{sa.sheet_name}"
                        f" vs {prev[0]} in {prev[1]}/{prev[2]}）"
                    )
                else:
                    seen[stu.student_no] = (stu.name, fa.file_name, sa.sheet_name)

    return report


# ============================================================
# 输出
# ============================================================

def print_report(report: AnalysisReport, verbose: bool = False):
    print("=" * 60)
    print("Excel 考勤表分析报告")
    print("=" * 60)
    print(f"扫描目录: {report.data_dir}")
    print(f"文件数: {report.total_files}")
    print(f"Sheet 数: {report.total_sheets}")
    print(f"学生总数: {report.total_students}")
    print()

    for fa in report.files:
        grade_str = f"{fa.grade_year}级" if fa.grade_year else "多年级"
        major_str = f"{fa.major_short}({fa.major_full})" if fa.major_short else "多专业"
        print(f"--- {fa.file_name} [{grade_str} / {major_str}] (格式: {fa.format_type}) ---")

        for sa in fa.sheets:
            if sa.format_type == "roster":
                # 名单格式: 统计班级分布
                class_counts = {}
                for stu in sa.students:
                    cl = stu.class_label or "未知班级"
                    class_counts[cl] = class_counts.get(cl, 0) + 1
                classes_str = ", ".join(f"{k}:{v}人" for k, v in sorted(class_counts.items()))
                print(f"  Sheet: {sa.sheet_name} (名单) | {len(sa.students)}人 | {classes_str}")
            else:
                code_str = sa.class_code or "未知"
                source_str = f"(来源: {sa.class_code_source})" if sa.class_code_source else "(待人工确认)"
                print(f"  Sheet: {sa.sheet_name} → {code_str} {source_str} | {len(sa.students)}人")

            if verbose and sa.students:
                for stu in sa.students:
                    cl = f" [{stu.class_label}]" if stu.class_label else ""
                    print(f"    Row {stu.row_number}: {stu.name} | {stu.student_no}{cl}")

        print()

    if report.warnings:
        print("=" * 60)
        print(f"警告 ({len(report.warnings)} 条):")
        print("=" * 60)
        for w in report.warnings:
            print(f"  [!] {w}")
        print()

    if report.errors:
        print("=" * 60)
        print(f"错误 ({len(report.errors)} 条):")
        print("=" * 60)
        for e in report.errors:
            print(f"  [X] {e}")
        print()


def main():
    parser = argparse.ArgumentParser(description="分析考勤表 Excel 文件结构")
    parser.add_argument("--dir", default=EXCEL_DATA_DIR, help="Excel 文件所在目录")
    parser.add_argument("--json", action="store_true", help="输出 JSON 格式")
    parser.add_argument("--verbose", action="store_true", help="显示每个学生的详细数据")
    args = parser.parse_args()

    report = scan_and_analyze(args.dir)

    if args.json:
        print(json.dumps(asdict(report), ensure_ascii=False, indent=2))
    else:
        print_report(report, verbose=args.verbose)

    if report.errors:
        sys.exit(1)


if __name__ == "__main__":
    main()
