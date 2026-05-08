"""
提交审核API
"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from typing import List, Dict, Any
from datetime import datetime
from io import BytesIO
from urllib.parse import quote
import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side

from app.core.database import get_db
from app.core.security import get_current_user
from app.models import User, Submission, SubmissionRecord, AttendanceRecord, AttendanceTask, Student, Class, Major, SubmissionSnapshot
from app.models.week import WeekExport
from app.schemas.submission import (
    CreateSubmissionRequest,
    SubmissionResponse,
    SubmissionDetailResponse,
    ApproveSubmissionRequest,
    RejectSubmissionRequest,
    WeekSummaryResponse,
    ExportStatusResponse,
    SubmissionRecordsResponse,
    RecordDetail,
    SubmissionAdminSearchResponse
)
from app.routers.week import get_current_week_config, calculate_week_number

router = APIRouter(prefix="/submissions", tags=["submissions"])


def _build_snapshot_data(db: Session, submission: Submission) -> dict:
    """为 approved submission 构建快照数据"""
    user = db.query(User).filter(User.id == submission.user_id).first()
    reviewer = None
    if submission.reviewer_id:
        reviewer = db.query(User).filter(User.id == submission.reviewer_id).first()
    
    submission_records = db.query(SubmissionRecord).filter(
        SubmissionRecord.submission_id == submission.id
    ).all()
    
    records = []
    for sr in submission_records:
        record = db.query(AttendanceRecord).filter(
            AttendanceRecord.id == sr.record_id
        ).first()
        if not record:
            continue
        
        student = db.query(Student).filter(Student.id == record.student_id).first()
        if not student:
            continue
        
        class_ = db.query(Class).filter(Class.id == student.class_id).first()
        major = None
        if class_:
            major = db.query(Major).filter(Major.id == class_.major_id).first()
        
        records.append({
            "record_id": record.id,
            "task_id": record.task_id,
            "student_id": student.id,
            "student_name": student.name,
            "student_no": student.student_no,
            "class_id": student.class_id,
            "class_name": class_.display_name if class_ else "未知",
            "major_short_name": major.short_name if major else "",
            "class_code": class_.class_code if class_ else "",
            "status": record.status,
            "remark": record.remark,
            "record_time": (record.updated_at or record.created_at).isoformat() if (record.updated_at or record.created_at) else None,
            "created_at": record.created_at.isoformat() if record.created_at else None,
            "updated_at": record.updated_at.isoformat() if record.updated_at else None,
        })
    
    return {
        "submission": {
            "id": submission.id,
            "user_id": submission.user_id,
            "week_number": submission.week_number,
            "status": submission.status,
            "class_names": submission.class_names,
            "submitted_at": submission.submitted_at.isoformat() if submission.submitted_at else None,
            "review_time": submission.review_time.isoformat() if submission.review_time else None,
            "review_note": submission.review_note,
        },
        "submitter": {
            "id": user.id if user else None,
            "real_name": user.real_name if user else None,
            "email": user.email if user else None,
        },
        "reviewer": {
            "id": reviewer.id if reviewer else None,
            "real_name": reviewer.real_name if reviewer else None,
            "email": reviewer.email if reviewer else None,
        },
        "records": records,
    }


def _build_record_from_db(db: Session, record) -> dict:
    """从数据库构建 record dict（fallback 用）"""
    student = db.query(Student).filter(Student.id == record.student_id).first()
    if not student:
        return None
    
    class_ = db.query(Class).filter(Class.id == student.class_id).first()
    major = None
    if class_:
        major = db.query(Major).filter(Major.id == class_.major_id).first()
    
    return {
        "record_id": record.id,
        "task_id": record.task_id,
        "student_id": student.id,
        "student_name": student.name,
        "student_no": student.student_no,
        "class_id": student.class_id,
        "class_name": class_.display_name if class_ else "未知",
        "major_short_name": major.short_name if major else "",
        "class_code": class_.class_code if class_ else "",
        "status": record.status,
        "remark": record.remark,
        "record_time": (record.updated_at or record.created_at).isoformat() if (record.updated_at or record.created_at) else None,
        "created_at": record.created_at.isoformat() if record.created_at else None,
        "updated_at": record.updated_at.isoformat() if record.updated_at else None,
    }


def _get_approved_records_from_snapshots(
    db: Session, week_number: int
) -> List[Dict[str, Any]]:
    """从快照获取 approved 记录，submission 级别 fallback"""
    approved_submissions = db.query(Submission).filter(
        Submission.week_number == week_number,
        Submission.status == "approved"
    ).all()
    
    if not approved_submissions:
        return []
    
    # 预加载该周所有 snapshots
    snapshots = db.query(SubmissionSnapshot).filter(
        SubmissionSnapshot.week_number == week_number
    ).all()
    snapshot_map = {s.submission_id: s for s in snapshots}
    
    all_records = []
    for sub in approved_submissions:
        snapshot = snapshot_map.get(sub.id)
        if snapshot:
            # 使用快照
            try:
                data = json.loads(snapshot.snapshot_data)
                for r in data.get("records", []):
                    all_records.append(r)
            except (json.JSONDecodeError, TypeError) as e:
                print(f"[ERROR] Failed to parse snapshot for submission {sub.id}: {e}")
                # fallback 到旧逻辑
                sub_records = _fallback_records_for_submission(db, sub.id)
                all_records.extend(sub_records)
        else:
            # fallback: 旧逻辑实时查询
            sub_records = _fallback_records_for_submission(db, sub.id)
            all_records.extend(sub_records)
    
    return all_records


def _fallback_records_for_submission(db: Session, submission_id: int) -> List[Dict[str, Any]]:
    """fallback: 实时查询单个 submission 的 records"""
    submission_records = db.query(SubmissionRecord).filter(
        SubmissionRecord.submission_id == submission_id
    ).all()
    
    records = []
    for sr in submission_records:
        record = db.query(AttendanceRecord).filter(
            AttendanceRecord.id == sr.record_id
        ).first()
        if not record:
            continue
        
        r = _build_record_from_db(db, record)
        if r:
            records.append(r)
    
    return records


@router.post("/", response_model=SubmissionResponse)
async def create_submission(
    body: CreateSubmissionRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """创建提交"""
    tasks = db.query(AttendanceTask).filter(
        AttendanceTask.id.in_(body.task_ids)
    ).all()
    
    # 校验：所有 task 必须存在
    found_task_ids = {t.id for t in tasks}
    missing_task_ids = set(body.task_ids) - found_task_ids
    if missing_task_ids:
        raise HTTPException(
            status_code=400,
            detail=f"任务不存在或已被删除: {', '.join(missing_task_ids)}"
        )
    
    for task in tasks:
        # 校验：任务必须属于当前用户
        if task.user_id is not None and task.user_id != current_user.id:
            raise HTTPException(
                status_code=403,
                detail=f"任务 {task.id} 不属于当前用户"
            )
        # 校验：任务状态必须是 completed
        if task.status != "completed":
            raise HTTPException(
                status_code=400,
                detail=f"任务 {task.id} 状态为 {task.status}，只有已完成的任务才能提交"
            )
    
    records = db.query(AttendanceRecord).filter(
        AttendanceRecord.task_id.in_(body.task_ids)
    ).all()

    # 按 task_id + student_id 去重，保留 updated_at 最新的一条
    deduped = {}
    for r in records:
        key = (r.task_id, r.student_id)
        if key not in deduped or (r.updated_at and deduped[key].updated_at and r.updated_at > deduped[key].updated_at):
            deduped[key] = r
    records = list(deduped.values())
    
    # 重复提交检查：只拦截关联了 pending/approved submission 的记录
    existing_submission = db.query(SubmissionRecord).join(Submission).filter(
        SubmissionRecord.record_id.in_([r.id for r in records if r.id]),
        Submission.status.in_(['pending', 'approved'])
    ).first()
    
    if existing_submission:
        raise HTTPException(
            status_code=400,
            detail="部分记录已提交，无法重复提交"
        )
    
    # 获取班级名称
    class_ids = set([r.class_id for r in records if r.class_id])
    class_names = ""
    if class_ids:
        classes = db.query(Class).filter(Class.id.in_(class_ids)).all()
        class_names = ", ".join([c.display_name for c in classes if c.display_name])
    
    submission = Submission(
        user_id=current_user.id,
        week_number=body.week_number,
        status="pending",
        class_names=class_names
    )
    db.add(submission)
    db.flush()
    
    for record in records:
        if record.id:
            sr = SubmissionRecord(
                submission_id=submission.id,
                record_id=record.id
            )
            db.add(sr)
    
    db.commit()
    db.refresh(submission)
    
    return submission


@router.get("/submitted-task-ids")
async def get_submitted_task_ids(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """获取已提交的任务ID列表（只包含 pending/approved 状态的提交）"""
    submissions = db.query(Submission).filter(
        Submission.user_id == current_user.id,
        Submission.status.in_(["pending", "approved"])
    ).all()
    
    submitted_task_ids = set()
    for sub in submissions:
        submission_records = db.query(SubmissionRecord).filter(
            SubmissionRecord.submission_id == sub.id
        ).all()
        for sr in submission_records:
            record = db.query(AttendanceRecord).filter(
                AttendanceRecord.id == sr.record_id
            ).first()
            if record:
                submitted_task_ids.add(record.task_id)
    
    return {"task_ids": list(submitted_task_ids)}


@router.get("/", response_model=List[SubmissionDetailResponse])
async def get_submissions(
    week_number: int = None,
    status: str = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """获取用户自己的提交列表"""
    query = db.query(Submission).filter(Submission.user_id == current_user.id)
    
    if week_number:
        query = query.filter(Submission.week_number == week_number)
    
    if status:
        query = query.filter(Submission.status == status)
    
    submissions = query.order_by(Submission.submitted_at.desc()).limit(100).all()
    
    result = []
    for sub in submissions:
        try:
            reviewer = None
            if sub.reviewer_id:
                reviewer = db.query(User).filter(User.id == sub.reviewer_id).first()
            
            submission_records = db.query(SubmissionRecord).filter(
                SubmissionRecord.submission_id == sub.id
            ).all()
            record_count = len(submission_records)
            
            result.append(SubmissionDetailResponse(
                id=sub.id,
                user_id=sub.user_id,
                user_name=current_user.real_name or current_user.email,
                user_email=current_user.email,
                week_number=sub.week_number,
                status=sub.status,
                reviewer_id=sub.reviewer_id,
                reviewer_name=reviewer.real_name if reviewer else None,
                review_time=sub.review_time,
                review_note=sub.review_note,
                submitted_at=sub.submitted_at,
                task_count=1,
                record_count=record_count,
                class_names=sub.class_names or ""
            ))
        except Exception as e:
            print(f"Error processing submission {sub.id}: {e}")
            continue
    
    return result


@router.get("/pending", response_model=List[SubmissionDetailResponse])
async def get_pending_submissions(
    week_number: int = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """获取待审核提交列表（管理员）"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    
    query = db.query(Submission).filter(Submission.status == "pending")
    
    if week_number:
        query = query.filter(Submission.week_number == week_number)
    
    submissions = query.all()
    
    result = []
    for sub in submissions:
        user = db.query(User).filter(User.id == sub.user_id).first()
        
        submission_records = db.query(SubmissionRecord).filter(
            SubmissionRecord.submission_id == sub.id
        ).all()
        record_count = len(submission_records)
        
        result.append(SubmissionDetailResponse(
            id=sub.id,
            user_id=sub.user_id,
            user_name=user.real_name if user else None,
            user_email=user.email if user else None,
            week_number=sub.week_number,
            status=sub.status,
            reviewer_id=sub.reviewer_id,
            reviewer_name=None,
            review_time=sub.review_time,
            review_note=sub.review_note,
            submitted_at=sub.submitted_at,
            task_count=1,
            record_count=record_count,
            class_names=sub.class_names or ""
        ))
    
    return result


@router.get("/reviewed", response_model=List[SubmissionDetailResponse])
async def get_reviewed_submissions(
    week_number: int = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """获取已审核提交列表（管理员）"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    
    query = db.query(Submission).filter(
        Submission.status.in_(["approved", "rejected"])
    )
    
    if week_number:
        query = query.filter(Submission.week_number == week_number)
    
    submissions = query.order_by(Submission.review_time.desc()).limit(100).all()
    
    result = []
    for sub in submissions:
        user = db.query(User).filter(User.id == sub.user_id).first()
        reviewer = None
        if sub.reviewer_id:
            reviewer = db.query(User).filter(User.id == sub.reviewer_id).first()
        
        submission_records = db.query(SubmissionRecord).filter(
            SubmissionRecord.submission_id == sub.id
        ).all()
        record_count = len(submission_records)
        
        result.append(SubmissionDetailResponse(
            id=sub.id,
            user_id=sub.user_id,
            user_name=user.real_name if user else None,
            user_email=user.email if user else None,
            week_number=sub.week_number,
            status=sub.status,
            reviewer_id=sub.reviewer_id,
            reviewer_name=reviewer.real_name if reviewer else None,
            review_time=sub.review_time,
            review_note=sub.review_note,
            submitted_at=sub.submitted_at,
            task_count=1,
            record_count=record_count,
            class_names=sub.class_names or ""
        ))
    
    return result


@router.get("/admin-search", response_model=SubmissionAdminSearchResponse)
async def admin_search_submissions(
    page: int = 1,
    page_size: int = 20,
    status: str = None,
    week_number: int = None,
    user_id: int = None,
    start_date: str = None,
    end_date: str = None,
    keyword: str = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """管理员查询提交记录（支持分页和筛选）"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    
    # 限制分页参数
    page = max(1, page)
    page_size = min(max(1, page_size), 100)
    
    query = db.query(Submission)
    
    # 状态筛选
    if status:
        query = query.filter(Submission.status == status)
    
    # 周次筛选
    if week_number:
        query = query.filter(Submission.week_number == week_number)
    
    # 提交人筛选
    if user_id:
        query = query.filter(Submission.user_id == user_id)
    
    # 日期范围筛选
    if start_date:
        try:
            start_dt = datetime.strptime(start_date, "%Y-%m-%d")
            query = query.filter(Submission.submitted_at >= start_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"开始日期格式错误: {start_date}，应为 YYYY-MM-DD")
    
    if end_date:
        try:
            end_dt = datetime.strptime(end_date, "%Y-%m-%d")
            # 包含当天结束
            end_dt = end_dt.replace(hour=23, minute=59, second=59)
            query = query.filter(Submission.submitted_at <= end_dt)
        except ValueError:
            raise HTTPException(status_code=400, detail=f"结束日期格式错误: {end_date}，应为 YYYY-MM-DD")
    
    # 关键词搜索（班级名称、提交人姓名/邮箱）
    if keyword:
        keyword = keyword.strip()
        # 先搜索用户
        matched_users = db.query(User).filter(
            (User.real_name.contains(keyword)) | 
            (User.email.contains(keyword))
        ).all()
        matched_user_ids = [u.id for u in matched_users]
        
        query = query.filter(
            (Submission.class_names.contains(keyword)) |
            (Submission.user_id.in_(matched_user_ids) if matched_user_ids else False)
        )
    
    # 统计总数
    total = query.count()
    
    # 分页
    offset = (page - 1) * page_size
    submissions = query.order_by(Submission.submitted_at.desc()).offset(offset).limit(page_size).all()
    
    result = []
    for sub in submissions:
        user = db.query(User).filter(User.id == sub.user_id).first()
        reviewer = None
        if sub.reviewer_id:
            reviewer = db.query(User).filter(User.id == sub.reviewer_id).first()
        
        submission_records = db.query(SubmissionRecord).filter(
            SubmissionRecord.submission_id == sub.id
        ).all()
        record_count = len(submission_records)
        
        # 统计关联的 task 数量（去重）
        task_ids = set()
        if submission_records:
            record_ids = [sr.record_id for sr in submission_records]
            records = db.query(AttendanceRecord).filter(
                AttendanceRecord.id.in_(record_ids)
            ).all()
            task_ids = {r.task_id for r in records if r.task_id}
        task_count = len(task_ids)
        
        result.append(SubmissionDetailResponse(
            id=sub.id,
            user_id=sub.user_id,
            user_name=user.real_name if user else None,
            user_email=user.email if user else None,
            week_number=sub.week_number,
            status=sub.status,
            reviewer_id=sub.reviewer_id,
            reviewer_name=reviewer.real_name if reviewer else None,
            review_time=sub.review_time,
            review_note=sub.review_note,
            submitted_at=sub.submitted_at,
            task_count=task_count,
            record_count=record_count,
            class_names=sub.class_names or ""
        ))
    
    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": result
    }


@router.get("/week-summary/{week_number}", response_model=WeekSummaryResponse)
async def get_week_summary(
    week_number: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """获取周汇总统计（优先使用快照）"""
    from sqlalchemy import func as sql_func
    
    submissions = db.query(Submission).filter(
        Submission.week_number == week_number
    ).all()
    
    pending = len([s for s in submissions if s.status == "pending"])
    approved = len([s for s in submissions if s.status == "approved"])
    rejected = len([s for s in submissions if s.status == "rejected"])
    
    late_count = 0
    absent_count = 0
    leave_count = 0
    other_count = 0
    late_student_count = 0
    absent_student_count = 0
    leave_student_count = 0
    other_student_count = 0
    total_abnormal_students = 0
    
    # 优先使用快照，fallback 到旧逻辑
    all_records = _get_approved_records_from_snapshots(db, week_number)
    
    # 从快照或 fallback 数据中统计
    late_students = set()
    absent_students = set()
    leave_students = set()
    other_students = set()
    
    for r in all_records:
        status = r.get("status")
        sid = r.get("student_id")
        if status == "late":
            late_count += 1
            late_students.add(sid)
        elif status == "absent":
            absent_count += 1
            absent_students.add(sid)
        elif status == "leave":
            leave_count += 1
            leave_students.add(sid)
        elif status == "other":
            other_count += 1
            other_students.add(sid)
    
    late_student_count = len(late_students)
    absent_student_count = len(absent_students)
    leave_student_count = len(leave_students)
    other_student_count = len(other_students)
    total_abnormal_students = len(late_students | absent_students | leave_students | other_students)
    
    export = db.query(WeekExport).filter(
        WeekExport.week_number == week_number
    ).order_by(WeekExport.exported_at.desc()).first()
    
    return WeekSummaryResponse(
        week_number=week_number,
        total_submissions=len(submissions),
        pending_count=pending,
        approved_count=approved,
        rejected_count=rejected,
        late_count=late_count,
        absent_count=absent_count,
        leave_count=leave_count,
        other_count=other_count,
        late_student_count=late_student_count,
        absent_student_count=absent_student_count,
        leave_student_count=leave_student_count,
        other_student_count=other_student_count,
        total_abnormal_students=total_abnormal_students,
        is_published=export is not None
    )


@router.get("/week-summary-detail/{week_number}")
async def get_week_summary_detail(
    week_number: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """获取周汇总详细名单（优先使用快照）"""
    export = db.query(WeekExport).filter(
        WeekExport.week_number == week_number
    ).first()
    
    if current_user.role != "admin" and not export:
        raise HTTPException(status_code=403, detail="该周汇总尚未发布")
    
    # 优先使用快照，fallback 到旧逻辑
    all_records = _get_approved_records_from_snapshots(db, week_number)
    
    student_stats = {}
    
    for r in all_records:
        status = r.get("status")
        if status not in ("late", "absent", "leave", "other"):
            continue
        
        sid = r.get("student_id")
        if sid not in student_stats:
            student_stats[sid] = {
                "student_id": sid,
                "name": r.get("student_name", ""),
                "student_no": r.get("student_no", ""),
                "class_name": r.get("class_name", "未知"),
                "major_short_name": r.get("major_short_name", ""),
                "class_code": r.get("class_code", ""),
                "late": 0,
                "absent": 0,
                "leave": 0,
                "other": 0,
            }
        
        if status == "late":
            student_stats[sid]["late"] += 1
        elif status == "absent":
            student_stats[sid]["absent"] += 1
        elif status == "leave":
            student_stats[sid]["leave"] += 1
        elif status == "other":
            student_stats[sid]["other"] += 1
    
    sorted_students = sorted(
        student_stats.values(),
        key=lambda x: (x["major_short_name"], int(x["class_code"]) if x["class_code"].isdigit() else 0, x["student_no"])
    )
    
    table_data = []
    for i, s in enumerate(sorted_students, 1):
        table_data.append({
            "index": i,
            "name": s["name"],
            "class_name": s["class_name"],
            "student_no": s["student_no"],
            "late": s["late"],
            "absent": s["absent"],
            "leave": s["leave"],
            "other": s["other"],
            "total": (s["late"] // 2) + s["absent"],
        })
    
    return {
        "week_number": week_number,
        "table_data": table_data,
        "total_count": len(table_data),
    }


@router.get("/history")
async def get_history_weeks(
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """获取历史周次列表（已导出的周次）"""
    exports = db.query(WeekExport).order_by(WeekExport.week_number.desc()).all()
    
    result = []
    for export in exports:
        result.append({
            "week_number": export.week_number,
            "exported_at": export.exported_at.isoformat() if export.exported_at else None,
        })
    
    return result


@router.get("/export-status/{week_number}", response_model=ExportStatusResponse)
async def get_export_status(
    week_number: int,
    db: Session = Depends(get_db)
):
    """获取周导出状态"""
    export = db.query(WeekExport).filter(
        WeekExport.week_number == week_number
    ).order_by(WeekExport.exported_at.desc()).first()
    
    if not export:
        return ExportStatusResponse(
            week_number=week_number,
            is_published=False
        )
    
    exporter = db.query(User).filter(User.id == export.exported_by).first()
    
    return ExportStatusResponse(
        week_number=week_number,
        is_published=True,
        exported_at=export.exported_at,
        exported_by_name=exporter.real_name if exporter else None
    )


@router.get("/export/{week_number}")
async def export_week_excel(
    week_number: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """导出周考勤Excel（管理员，优先使用快照）"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    
    # 优先使用快照，fallback 到旧逻辑
    all_records = _get_approved_records_from_snapshots(db, week_number)
    
    if not all_records:
        raise HTTPException(status_code=400, detail="该周无已审核通过的提交")
    
    student_stats = {}
    for r in all_records:
        status = r.get("status")
        if status not in ("late", "absent", "leave", "other"):
            continue
        
        sid = r.get("student_id")
        if sid not in student_stats:
            student_stats[sid] = {
                "name": r.get("student_name", ""),
                "student_no": r.get("student_no", ""),
                "class_name": r.get("class_name", "未知"),
                "major_short_name": r.get("major_short_name", ""),
                "class_code": r.get("class_code", ""),
                "late": 0,
                "absent": 0,
                "leave": 0,
                "other": 0,
                "record_times": [],
            }
        
        if status == "late":
            student_stats[sid]["late"] += 1
        elif status == "absent":
            student_stats[sid]["absent"] += 1
        elif status == "leave":
            student_stats[sid]["leave"] += 1
        elif status == "other":
            student_stats[sid]["other"] += 1
        
        # 记录时间和状态
        record_time_str = r.get("record_time")
        if record_time_str:
            try:
                record_time = datetime.fromisoformat(record_time_str)
                status_label = {
                    "late": "迟",
                    "absent": "缺",
                    "leave": "假",
                    "other": "其"
                }.get(status, "")
                student_stats[sid]["record_times"].append({
                    "time": record_time,
                    "status_label": status_label
                })
            except (ValueError, TypeError):
                pass
    
    sorted_students = sorted(
        student_stats.values(),
        key=lambda x: (x["major_short_name"], int(x["class_code"]) if x["class_code"].isdigit() else 0, x["student_no"])
    )
    
    filtered_students = [
        s for s in sorted_students
        if s["late"] > 0 or s["absent"] > 0
    ]
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"第{week_number}周考勤"
    
    headers = ["序号", "姓名", "班级", "学号", "迟到", "缺勤", "累计", "记录时间"]
    ws.append(headers)
    
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    
    for i, s in enumerate(filtered_students, 1):
        # 格式化时间记录
        time_str = ""
        if s["record_times"]:
            # 按时间排序
            sorted_times = sorted(s["record_times"], key=lambda x: x["time"])
            time_parts = []
            for rt in sorted_times:
                t = rt["time"]
                formatted = f"{t.month:02d}-{t.day:02d} {t.hour:02d}:{t.minute:02d}({rt['status_label']})"
                time_parts.append(formatted)
            time_str = "\n".join(time_parts)
        
        row = [
            i,
            s["name"],
            s["class_name"],
            s["student_no"],
            s["late"],
            s["absent"],
            None,
            time_str,
        ]
        ws.append(row)
        
        ws.cell(row=i+1, column=7).value = f"=ROUNDDOWN(E{i+1}/2+F{i+1},0)"
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=8):
        for cell in row:
            cell.border = thin_border
            # 时间列(H列)保持左对齐，其他列居中
            if cell.column == 8:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center")
    
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 8
    ws.column_dimensions['F'].width = 8
    ws.column_dimensions['G'].width = 8
    ws.column_dimensions['H'].width = 20
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    export = WeekExport(
        week_number=week_number,
        exported_by=current_user.id
    )
    db.add(export)
    db.commit()
    
    filename = f"第{week_number}周考勤表.xlsx"
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}"}
    )


@router.get("/{submission_id}", response_model=SubmissionDetailResponse)
async def get_submission_detail(
    submission_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """获取提交详情"""
    submission = db.query(Submission).filter(Submission.id == submission_id).first()
    
    if not submission:
        raise HTTPException(status_code=404, detail="提交不存在")
    
    if submission.user_id != current_user.id and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="无权查看此提交")
    
    user = db.query(User).filter(User.id == submission.user_id).first()
    reviewer = None
    if submission.reviewer_id:
        reviewer = db.query(User).filter(User.id == submission.reviewer_id).first()
    
    submission_records = db.query(SubmissionRecord).filter(
        SubmissionRecord.submission_id == submission.id
    ).all()
    record_count = len(submission_records)
    
    class_names = _get_submission_class_names(db, submission.id)
    
    return SubmissionDetailResponse(
        id=submission.id,
        user_id=submission.user_id,
        user_name=user.real_name if user else None,
        user_email=user.email if user else None,
        week_number=submission.week_number,
        status=submission.status,
        reviewer_id=submission.reviewer_id,
        reviewer_name=reviewer.real_name if reviewer else None,
        review_time=submission.review_time,
        review_note=submission.review_note,
        submitted_at=submission.submitted_at,
        task_count=1,
        record_count=record_count,
        class_names=class_names
    )


@router.get("/{submission_id}/records", response_model=SubmissionRecordsResponse)
async def get_submission_records(
    submission_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """获取提交的详细记录列表"""
    submission = db.query(Submission).filter(Submission.id == submission_id).first()
    
    if not submission:
        raise HTTPException(status_code=404, detail="提交不存在")
    
    if submission.user_id != current_user.id and current_user.role != "admin":
        raise HTTPException(status_code=403, detail="无权查看此提交")
    
    user = db.query(User).filter(User.id == submission.user_id).first()
    
    submission_records = db.query(SubmissionRecord).filter(
        SubmissionRecord.submission_id == submission.id
    ).all()
    
    records = []
    late_count = 0
    absent_count = 0
    leave_count = 0
    other_count = 0
    
    for sr in submission_records:
        record = db.query(AttendanceRecord).filter(
            AttendanceRecord.id == sr.record_id
        ).first()
        
        if record:
            student = db.query(Student).filter(
                Student.id == record.student_id
            ).first()
            
            if student:
                class_ = db.query(Class).filter(Class.id == student.class_id).first()
                
                records.append(RecordDetail(
                    student_id=student.id,
                    student_name=student.name,
                    student_no=student.student_no,
                    class_name=class_.display_name if class_ else "未知",
                    status=record.status
                ))
                
                if record.status == "late":
                    late_count += 1
                elif record.status == "absent":
                    absent_count += 1
                elif record.status == "leave":
                    leave_count += 1
                elif record.status == "other":
                    other_count += 1
    
    return SubmissionRecordsResponse(
        id=submission.id,
        user_id=submission.user_id,
        user_name=user.real_name if user else None,
        user_email=user.email if user else None,
        week_number=submission.week_number,
        status=submission.status,
        submitted_at=submission.submitted_at,
        records=records,
        late_count=late_count,
        absent_count=absent_count,
        leave_count=leave_count,
        other_count=other_count
    )


@router.put("/{submission_id}/approve")
async def approve_submission(
    submission_id: int,
    body: ApproveSubmissionRequest = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """审核通过"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    
    submission = db.query(Submission).filter(Submission.id == submission_id).first()
    
    if not submission:
        raise HTTPException(status_code=404, detail="提交不存在")
    
    if submission.status == "cancelled":
        raise HTTPException(status_code=400, detail="成员已撤销该提交，无法审核")
    
    if submission.status != "pending":
        status_text = {
            "approved": "已通过",
            "rejected": "已拒绝",
        }.get(submission.status, submission.status)
        raise HTTPException(status_code=400, detail=f"提交状态不是待审核（当前状态：{status_text}）")
    
    submission.status = "approved"
    submission.reviewer_id = current_user.id
    submission.review_time = datetime.now()
    if body and body.note:
        submission.review_note = body.note
    
    # 生成快照（幂等：如果已存在则不重复创建）
    existing_snapshot = db.query(SubmissionSnapshot).filter(
        SubmissionSnapshot.submission_id == submission_id
    ).first()
    
    if not existing_snapshot:
        snapshot_data = _build_snapshot_data(db, submission)
        snapshot = SubmissionSnapshot(
            submission_id=submission.id,
            week_number=submission.week_number,
            user_id=submission.user_id,
            class_names=submission.class_names,
            snapshot_data=json.dumps(snapshot_data, ensure_ascii=False)
        )
        db.add(snapshot)
    
    db.commit()
    
    return {"message": "审核通过", "submission_id": submission_id}


@router.put("/{submission_id}/reject")
async def reject_submission(
    submission_id: int,
    body: RejectSubmissionRequest,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """审核拒绝"""
    if current_user.role != "admin":
        raise HTTPException(status_code=403, detail="需要管理员权限")
    
    submission = db.query(Submission).filter(Submission.id == submission_id).first()
    
    if not submission:
        raise HTTPException(status_code=404, detail="提交不存在")
    
    if submission.status == "cancelled":
        raise HTTPException(status_code=400, detail="成员已撤销该提交，无法审核")
    
    if submission.status != "pending":
        status_text = {
            "approved": "已通过",
            "rejected": "已拒绝",
        }.get(submission.status, submission.status)
        raise HTTPException(status_code=400, detail=f"提交状态不是待审核（当前状态：{status_text}）")
    
    submission.status = "rejected"
    submission.reviewer_id = current_user.id
    submission.review_time = datetime.now()
    submission.review_note = body.note
    
    db.query(SubmissionRecord).filter(
        SubmissionRecord.submission_id == submission_id
    ).delete()
    
    db.commit()
    
    return {"message": "审核拒绝", "submission_id": submission_id}


@router.delete("/{submission_id}")
async def cancel_submission(
    submission_id: int,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """撤销提交"""
    submission = db.query(Submission).filter(Submission.id == submission_id).first()
    
    if not submission:
        raise HTTPException(status_code=404, detail="提交不存在")
    
    if submission.user_id != current_user.id:
        raise HTTPException(status_code=403, detail="只能撤销自己的提交")
    
    if submission.status == "cancelled":
        raise HTTPException(status_code=400, detail="该提交已被撤销")
    
    if submission.status == "approved":
        raise HTTPException(status_code=400, detail="管理员已审核通过，无法撤销")
    
    if submission.status == "rejected":
        raise HTTPException(status_code=400, detail="管理员已拒绝该提交，无法撤销")
    
    if submission.status != "pending":
        raise HTTPException(status_code=400, detail="只能撤销待审核的提交")
    
    db.query(SubmissionRecord).filter(
        SubmissionRecord.submission_id == submission_id
    ).delete()
    
    submission.status = "cancelled"
    db.commit()
    
    return {"message": "提交已撤销", "submission_id": submission_id}